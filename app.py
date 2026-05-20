"""
=============================================================
  SISTEM SELEKSI OLIMPIADE MATEMATIKA
  File: app/app.py
  Deskripsi: Aplikasi Streamlit untuk prediksi kesiapan
             olimpiade matematika siswa berbasis ML
=============================================================
"""

import streamlit as st
import pandas as pd
import numpy as np
import joblib
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO
import os
import sys

# ─── Path Config ────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MODEL_DIR = os.path.join(BASE_DIR, "model")

# ─── Page Config ─────────────────────────────────────────────
st.set_page_config(
    page_title="OlympiQ — Seleksi Olimpiade Matematika",
    page_icon="🏆",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── CSS Styling ─────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=Syne:wght@400;600;700;800&display=swap');

  html, body, [class*="css"] {
    font-family: 'Space Grotesk', sans-serif;
  }

  .main { background: #0d0f1a; }
  .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

  /* Header Hero */
  .hero-box {
    background: linear-gradient(135deg, #1a1f35 0%, #0d1226 50%, #0a0d1a 100%);
    border: 1px solid rgba(99,179,237,0.2);
    border-radius: 20px;
    padding: 2.5rem 3rem;
    margin-bottom: 1.5rem;
    position: relative;
    overflow: hidden;
  }
  .hero-box::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 500px;
    height: 500px;
    background: radial-gradient(circle, rgba(99,179,237,0.08) 0%, transparent 70%);
    pointer-events: none;
  }
  .hero-title {
    font-family: 'Syne', sans-serif;
    font-size: 2.6rem;
    font-weight: 800;
    color: #ffffff;
    margin: 0 0 0.4rem 0;
    letter-spacing: -1px;
  }
  .hero-title span { color: #63b3ed; }
  .hero-sub {
    font-size: 1rem;
    color: rgba(255,255,255,0.55);
    margin: 0;
    font-weight: 400;
  }
  .hero-badge {
    display: inline-block;
    background: rgba(99,179,237,0.15);
    border: 1px solid rgba(99,179,237,0.3);
    color: #63b3ed;
    font-size: 0.72rem;
    font-weight: 600;
    padding: 0.3rem 0.8rem;
    border-radius: 50px;
    margin-bottom: 1rem;
    letter-spacing: 1.5px;
    text-transform: uppercase;
  }

  /* Cards */
  .card {
    background: #12172b;
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 16px;
    padding: 1.5rem;
    margin-bottom: 1rem;
  }
  .card-title {
    font-family: 'Syne', sans-serif;
    font-size: 1.1rem;
    font-weight: 700;
    color: #ffffff;
    margin: 0 0 1rem 0;
  }

  /* Status Badges */
  .status-siap {
    background: linear-gradient(135deg, #065f46, #047857);
    border: 1px solid #10b981;
    color: #6ee7b7;
    padding: 0.9rem 1.8rem;
    border-radius: 50px;
    font-size: 1.1rem;
    font-weight: 700;
    display: inline-block;
    letter-spacing: 0.5px;
  }
  .status-potensial {
    background: linear-gradient(135deg, #78350f, #92400e);
    border: 1px solid #f59e0b;
    color: #fde68a;
    padding: 0.9rem 1.8rem;
    border-radius: 50px;
    font-size: 1.1rem;
    font-weight: 700;
    display: inline-block;
    letter-spacing: 0.5px;
  }
  .status-tidak {
    background: linear-gradient(135deg, #7f1d1d, #991b1b);
    border: 1px solid #ef4444;
    color: #fca5a5;
    padding: 0.9rem 1.8rem;
    border-radius: 50px;
    font-size: 1.1rem;
    font-weight: 700;
    display: inline-block;
    letter-spacing: 0.5px;
  }

  /* Review & Rekomendasi */
  .review-box {
    background: rgba(99,179,237,0.06);
    border-left: 3px solid #63b3ed;
    border-radius: 0 12px 12px 0;
    padding: 1rem 1.2rem;
    margin: 0.5rem 0;
    color: rgba(255,255,255,0.85);
    font-size: 0.92rem;
    line-height: 1.7;
  }
  .saran-item {
    background: rgba(255,255,255,0.04);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 10px;
    padding: 0.75rem 1rem;
    margin: 0.4rem 0;
    color: rgba(255,255,255,0.8);
    font-size: 0.9rem;
  }
  .saran-item::before { content: "→  "; color: #63b3ed; font-weight: 700; }

  /* Metric Cards */
  .metric-grid { display: flex; gap: 0.8rem; flex-wrap: wrap; margin-bottom: 0.8rem; }
  .metric-card {
    background: #1a2040;
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 12px;
    padding: 0.8rem 1rem;
    flex: 1;
    min-width: 120px;
    text-align: center;
  }
  .metric-label { font-size: 0.7rem; color: rgba(255,255,255,0.45); text-transform: uppercase; letter-spacing: 1px; margin-bottom: 0.3rem; }
  .metric-value { font-size: 1.4rem; font-weight: 700; color: #63b3ed; }

  /* Input Styling */
  .stSlider > div > div { color: #63b3ed; }
  .stNumberInput input { background: #1a2040; border: 1px solid rgba(99,179,237,0.3); border-radius: 8px; color: #fff; }

  /* Tab */
  .stTabs [data-baseweb="tab"] {
    background: #12172b;
    border: 1px solid rgba(255,255,255,0.06);
    border-radius: 8px 8px 0 0;
    color: rgba(255,255,255,0.5);
    font-weight: 500;
  }
  .stTabs [aria-selected="true"] {
    background: #1a2040 !important;
    color: #63b3ed !important;
    border-bottom: 2px solid #63b3ed !important;
  }

  /* Button */
  .stButton > button {
    background: linear-gradient(135deg, #2563eb, #3b82f6);
    color: white;
    border: none;
    border-radius: 10px;
    font-weight: 600;
    font-size: 0.95rem;
    padding: 0.6rem 2rem;
    width: 100%;
    transition: all 0.2s;
  }
  .stButton > button:hover {
    background: linear-gradient(135deg, #1d4ed8, #2563eb);
    transform: translateY(-1px);
    box-shadow: 0 8px 25px rgba(37,99,235,0.4);
  }

  /* Sidebar */
  [data-testid="stSidebar"] {
    background: #0d1226;
    border-right: 1px solid rgba(255,255,255,0.06);
  }
  [data-testid="stSidebar"] .stMarkdown p { color: rgba(255,255,255,0.6); font-size: 0.85rem; }

  /* Section Header */
  .section-label {
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: #63b3ed;
    margin: 1.2rem 0 0.5rem 0;
  }

  /* Probability bars */
  .prob-row { display: flex; align-items: center; gap: 0.8rem; margin: 0.4rem 0; }
  .prob-label { width: 120px; font-size: 0.82rem; color: rgba(255,255,255,0.65); }
  .prob-bar-bg { flex: 1; background: rgba(255,255,255,0.07); border-radius: 50px; height: 8px; overflow: hidden; }
  .prob-fill-green  { height: 100%; border-radius: 50px; background: linear-gradient(90deg, #10b981, #34d399); }
  .prob-fill-yellow { height: 100%; border-radius: 50px; background: linear-gradient(90deg, #f59e0b, #fbbf24); }
  .prob-fill-red    { height: 100%; border-radius: 50px; background: linear-gradient(90deg, #ef4444, #f87171); }
  .prob-pct { width: 45px; font-size: 0.82rem; color: rgba(255,255,255,0.7); text-align: right; font-weight: 600; }

  /* DataTable */
  .dataframe { background: #12172b !important; }
  .stDataFrame { border-radius: 12px; overflow: hidden; }

  div[data-testid="metric-container"] {
    background: #12172b;
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 12px;
    padding: 0.8rem;
  }
  div[data-testid="metric-container"] label { color: rgba(255,255,255,0.5) !important; font-size: 0.75rem !important; }
  div[data-testid="metric-container"] [data-testid="metric-value"] { color: #63b3ed !important; font-size: 1.6rem !important; font-weight: 700 !important; }

  .divider { border: none; border-top: 1px solid rgba(255,255,255,0.07); margin: 1.2rem 0; }
</style>
""", unsafe_allow_html=True)


# ─── Load Model ─────────────────────────────────────────────
@st.cache_resource
def load_model():
    rf        = joblib.load(os.path.join(MODEL_DIR, "rf_model.pkl"))
    scaler    = joblib.load(os.path.join(MODEL_DIR, "scaler.pkl"))
    thresholds = joblib.load(os.path.join(MODEL_DIR, "thresholds.pkl"))
    meta      = joblib.load(os.path.join(MODEL_DIR, "model_meta.pkl"))
    return rf, scaler, thresholds, meta

try:
    rf_model, scaler, thresholds, meta = load_model()
    MODEL_LOADED = True
except Exception as e:
    MODEL_LOADED = False
    st.error(f"❌ Model tidak ditemukan. Jalankan `python train_model.py` terlebih dahulu.\nError: {e}")
    st.stop()


# ─── Konstanta ──────────────────────────────────────────────
FEATURE_COLS  = ['NUM_ALJ', 'NUM_GEO', 'NUM_BIL', 'NUM_DAT', 'NUM_L3', 'LIT']
FEATURE_LABELS = {
    'NUM_ALJ': 'Numerasi Aljabar',
    'NUM_GEO': 'Numerasi Geometri',
    'NUM_BIL': 'Numerasi Bilangan',
    'NUM_DAT': 'Data & Ketidakpastian',
    'NUM_L3':  'Skor Menalar',
    'LIT':     'Literasi'
}
CLASS_MAP = {0: 'Tidak Siap', 1: 'Potensial', 2: 'Siap Olimpiade'}
STATUS_EMOJI = {0: '❌', 1: '⚡', 2: '🏆'}
WEIGHTS = {'NUM_ALJ': 0.20, 'NUM_GEO': 0.15, 'NUM_BIL': 0.20,
           'NUM_DAT': 0.15, 'NUM_L3': 0.15, 'LIT': 0.15}

# ─── Helper Functions ────────────────────────────────────────
def predict_one(scores: dict):
    X = pd.DataFrame([scores])[FEATURE_COLS]
    X_sc = scaler.transform(X)
    label  = rf_model.predict(X_sc)[0]
    proba  = rf_model.predict_proba(X_sc)[0]
    total  = sum(scores[c] * WEIGHTS[c] for c in FEATURE_COLS)
    return label, proba, total

def get_review(label, scores, total):
    weak  = [FEATURE_LABELS[c] for c in FEATURE_COLS if scores[c] < 55]
    strong = [FEATURE_LABELS[c] for c in FEATURE_COLS if scores[c] >= 80]
    if label == 2:
        review = (
            f"Siswa ini menunjukkan performa sangat unggul dengan skor komprehensif {total:.1f}/100. "
            f"Kemampuan multidimensi yang merata mencerminkan kesiapan tinggi untuk kompetisi olimpiade matematika. "
        )
        if strong:
            review += f"Keunggulan menonjol pada: {', '.join(strong)}."
    elif label == 1:
        review = (
            f"Siswa ini memiliki potensi yang cukup kuat dengan skor {total:.1f}/100. "
            f"Beberapa aspek sudah cukup matang namun masih ada ruang untuk peningkatan. "
        )
        if weak:
            review += f"Area yang perlu diperhatikan: {', '.join(weak)}."
    else:
        review = (
            f"Skor komprehensif siswa ({total:.1f}/100) menunjukkan bahwa kompetensi dasar masih perlu diperkuat "
            f"sebelum mengikuti seleksi olimpiade. "
        )
        if weak:
            review += f"Fokus prioritas: {', '.join(weak)}."
    return review

def get_recommendations(label, scores):
    recs = []
    if label == 2:
        recs = [
            "Daftarkan siswa ke program pembinaan olimpiade intensif tingkat lanjut",
            "Tingkatkan pengalaman kompetisi dengan mengikuti simulasi soal olimpiade nasional",
            "Berikan latihan soal level olimpiade internasional (IMO, APMO) untuk memperluas wawasan",
            "Perkuat kemampuan matematika kreatif dan pembuktian formal",
            "Rekomendasikan ke guru pembina untuk masuk tim inti olimpiade sekolah"
        ]
    elif label == 1:
        weak_areas = [FEATURE_LABELS[c] for c in FEATURE_COLS if scores[c] < 65]
        recs = [
            "Ikutkan dalam program pembinaan olimpiade reguler selama minimal 3 bulan ke depan",
            f"Fokus intensif pada: {', '.join(weak_areas) if weak_areas else 'semua aspek secara merata'}",
            "Berikan latihan soal olimpiade tingkat kabupaten/kota sebagai target awal",
            "Evaluasi ulang kesiapan setelah program pembinaan selesai dilaksanakan",
            "Tingkatkan kemampuan analisis melalui studi kasus dan problem solving terbimbing"
        ]
    else:
        weak_areas = [FEATURE_LABELS[c] for c in FEATURE_COLS if scores[c] < 60]
        recs = [
            "Prioritaskan penguatan kompetensi dasar matematika terlebih dahulu",
            f"Program remediasi khusus untuk: {', '.join(weak_areas) if weak_areas else 'semua komponen'}",
            "Gunakan metode pembelajaran diferensiasi sesuai gaya belajar siswa",
            "Pantau perkembangan secara berkala melalui asesmen formatif mingguan",
            "Pertimbangkan program pendampingan belajar dari teman sebaya (peer tutoring)"
        ]
    return recs

def get_saran(label, scores):
    saran = []
    for col, label_name in FEATURE_LABELS.items():
        s = scores[col]
        if s < 50:
            saran.append(f"[{label_name}] Skor rendah ({s:.1f}). Mulai dari materi dasar dan perbanyak latihan soal rutin.")
        elif s < 65:
            saran.append(f"[{label_name}] Skor cukup ({s:.1f}). Tingkatkan dengan latihan soal variasi dan diskusi kelompok.")
        elif s < 80:
            saran.append(f"[{label_name}] Skor baik ({s:.1f}). Pertahankan dan coba tantangan soal level lebih tinggi.")
        else:
            saran.append(f"[{label_name}] Skor sangat baik ({s:.1f}). Kuasai sudah sangat mantap, terus asah ke level olimpiade.")
    return saran

def radar_chart(scores, title="Profil Kompetensi Siswa"):
    cats = [FEATURE_LABELS[c] for c in FEATURE_COLS]
    vals = [scores[c] for c in FEATURE_COLS]
    vals_closed = vals + [vals[0]]
    cats_closed  = cats + [cats[0]]
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=vals_closed, theta=cats_closed, fill='toself',
        fillcolor='rgba(99,179,237,0.15)',
        line=dict(color='#63b3ed', width=2.5),
        name='Skor Siswa'
    ))
    fig.add_trace(go.Scatterpolar(
        r=[70]*7, theta=cats_closed, fill=None,
        line=dict(color='rgba(245,158,11,0.4)', width=1.5, dash='dot'),
        name='Target Min. (70)'
    ))
    fig.update_layout(
        polar=dict(
            radialaxis=dict(visible=True, range=[0, 100],
                            tickfont=dict(color='rgba(255,255,255,0.4)', size=9),
                            gridcolor='rgba(255,255,255,0.08)',
                            linecolor='rgba(255,255,255,0.1)'),
            angularaxis=dict(tickfont=dict(color='rgba(255,255,255,0.7)', size=10),
                             gridcolor='rgba(255,255,255,0.06)',
                             linecolor='rgba(255,255,255,0.06)')
        ),
        showlegend=True,
        legend=dict(font=dict(color='rgba(255,255,255,0.6)', size=10)),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=60, r=60, t=30, b=30),
        title=dict(text=title, font=dict(color='rgba(255,255,255,0.7)', size=12), x=0.5),
        height=360
    )
    return fig

def bar_chart(scores):
    labels = [FEATURE_LABELS[c] for c in FEATURE_COLS]
    values = [scores[c] for c in FEATURE_COLS]
    colors = ['#10b981' if v >= 75 else '#f59e0b' if v >= 55 else '#ef4444' for v in values]
    fig = go.Figure(go.Bar(
        x=labels, y=values,
        marker_color=colors,
        text=[f"{v:.1f}" for v in values],
        textposition='outside',
        textfont=dict(color='rgba(255,255,255,0.8)', size=11)
    ))
    fig.add_hline(y=70, line_dash="dot", line_color="rgba(245,158,11,0.5)",
                  annotation_text="Target 70", annotation_font_color="rgba(245,158,11,0.8)")
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(tickfont=dict(color='rgba(255,255,255,0.6)', size=10),
                   gridcolor='rgba(0,0,0,0)', showline=False),
        yaxis=dict(range=[0, 110], tickfont=dict(color='rgba(255,255,255,0.4)', size=9),
                   gridcolor='rgba(255,255,255,0.05)'),
        margin=dict(l=20, r=20, t=20, b=20),
        height=300
    )
    return fig

def to_excel_download(df_out: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Hasil Prediksi
        df_out.to_excel(writer, sheet_name='Hasil Prediksi', index=False)
        ws = writer.sheets['Hasil Prediksi']

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Header styling
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=10, name="Calibri")

        # Color status column
        status_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Status':
                status_col = idx
                break

        if status_col:
            for row in ws.iter_rows(min_row=2):
                cell = row[status_col - 1]
                if cell.value == 'Siap Olimpiade':
                    cell.fill = PatternFill("solid", fgColor="D4EDDA")
                    cell.font = Font(color="155724", bold=True, size=10, name="Calibri")
                elif cell.value == 'Potensial':
                    cell.fill = PatternFill("solid", fgColor="FFF3CD")
                    cell.font = Font(color="856404", bold=True, size=10, name="Calibri")
                elif cell.value == 'Tidak Siap':
                    cell.fill = PatternFill("solid", fgColor="F8D7DA")
                    cell.font = Font(color="721C24", bold=True, size=10, name="Calibri")

        # Column widths
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(max_len + 3, 35))

        # Sheet 2: Ringkasan
        if 'Status' in df_out.columns:
            summary = df_out['Status'].value_counts().reset_index()
            summary.columns = ['Status', 'Jumlah Siswa']
            summary['Persentase (%)'] = (summary['Jumlah Siswa'] / len(df_out) * 100).round(2)
            summary.to_excel(writer, sheet_name='Ringkasan', index=False)
            ws2 = writer.sheets['Ringkasan']
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

    return output.getvalue()


# ─── SIDEBAR ─────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='text-align:center; padding: 1rem 0;'>
      <div style='font-family:Syne,sans-serif; font-size:1.3rem; font-weight:800; color:#63b3ed;'>🏆 OlympiQ</div>
      <div style='font-size:0.72rem; color:rgba(255,255,255,0.35); letter-spacing:1.5px; text-transform:uppercase;'>Seleksi Olimpiade Matematika</div>
    </div>
    <hr style='border-color:rgba(255,255,255,0.07);'>
    """, unsafe_allow_html=True)

    st.markdown("<div class='section-label'>📊 Statistik Model</div>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    col1.metric("Akurasi", f"{meta['accuracy']*100:.1f}%")
    col2.metric("CV Score", f"{meta['cv_mean']*100:.1f}%")
    col1.metric("Data Latih", f"{meta['n_train']//1000}K")
    col2.metric("Data Uji", f"{meta['n_test']//1000}K")

    st.markdown("<hr style='border-color:rgba(255,255,255,0.07);'>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>🔍 Tentang Sistem</div>", unsafe_allow_html=True)
    st.markdown("""
    <p>Sistem ini menggunakan algoritma <b>Random Forest Classifier</b> yang dilatih pada <b>29.559 data siswa</b> untuk mengklasifikasikan kesiapan olimpiade ke dalam tiga kategori.</p>
    <p>Fitur yang digunakan:</p>
    """, unsafe_allow_html=True)

    for feat, label in FEATURE_LABELS.items():
        fi = meta['feature_importance'][feat]
        st.markdown(f"""
        <div style='display:flex; align-items:center; gap:0.5rem; margin:0.2rem 0;'>
          <div style='font-size:0.8rem; color:rgba(255,255,255,0.6); flex:1;'>{label}</div>
          <div style='font-size:0.75rem; color:#63b3ed; font-weight:600;'>{fi*100:.1f}%</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<hr style='border-color:rgba(255,255,255,0.07);'>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>📋 Keterangan Status</div>", unsafe_allow_html=True)
    st.markdown("""
    <div style='font-size:0.82rem; color:rgba(255,255,255,0.6); line-height:1.8;'>
    🏆 <b style='color:#6ee7b7;'>Siap Olimpiade</b><br>
    &nbsp;&nbsp;Kompetensi tinggi, siap bersaing<br><br>
    ⚡ <b style='color:#fde68a;'>Potensial</b><br>
    &nbsp;&nbsp;Kompetensi sedang, perlu pembinaan<br><br>
    ❌ <b style='color:#fca5a5;'>Tidak Siap</b><br>
    &nbsp;&nbsp;Perlu penguatan dasar terlebih dahulu
    </div>
    """, unsafe_allow_html=True)


# ─── HERO HEADER ─────────────────────────────────────────────
st.markdown("""
<div class='hero-box'>
  <div class='hero-badge'>🔬 Machine Learning · Random Forest Classifier</div>
  <div class='hero-title'><span>OlympiQ</span> — Sistem Seleksi Tahap Awal</div>
  <div class='hero-title' style='font-size:1.4rem; color:rgba(255,255,255,0.7); font-weight:500;'>Olimpiade Matematika Siswa</div>
  <div class='hero-sub' style='margin-top:0.8rem;'>Klasifikasi berbasis AI untuk membantu guru mengidentifikasi siswa berpotensi olimpiade secara akurat, efisien, dan transparan.</div>
</div>
""", unsafe_allow_html=True)


# ─── MAIN TABS ───────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["✏️ Input Satu Siswa", "📂 Upload File Excel", "📈 Informasi Model"])


# ═══════════════════════════════════════════════════════════
#  TAB 1 — INPUT SATU SISWA
# ═══════════════════════════════════════════════════════════
with tab1:
    st.markdown("<div class='card-title' style='margin-top:0.5rem;'>📝 Masukkan Skor Siswa</div>", unsafe_allow_html=True)
    st.markdown("<p style='color:rgba(255,255,255,0.45); font-size:0.85rem; margin-bottom:1rem;'>Isi seluruh skor di bawah (skala 0–100), kemudian klik tombol Prediksi.</p>", unsafe_allow_html=True)

    col_name, _ = st.columns([2, 3])
    with col_name:
        nama_siswa = st.text_input("Nama Siswa", placeholder="Contoh: Budi Santoso", label_visibility="visible")

    col1, col2, col3 = st.columns(3)
    with col1:
        num_alj = st.number_input("Numerasi Aljabar",     min_value=0.0, max_value=100.0, value=65.0, step=0.5)
        num_geo = st.number_input("Numerasi Geometri",    min_value=0.0, max_value=100.0, value=65.0, step=0.5)
    with col2:
        num_bil = st.number_input("Numerasi Bilangan",    min_value=0.0, max_value=100.0, value=65.0, step=0.5)
        num_dat = st.number_input("Data & Ketidakpastian",min_value=0.0, max_value=100.0, value=65.0, step=0.5)
    with col3:
        num_l3  = st.number_input("Skor Menalar",         min_value=0.0, max_value=100.0, value=65.0, step=0.5)
        lit     = st.number_input("Literasi",             min_value=0.0, max_value=100.0, value=72.0, step=0.5)

    scores = {'NUM_ALJ': num_alj, 'NUM_GEO': num_geo, 'NUM_BIL': num_bil,
              'NUM_DAT': num_dat, 'NUM_L3': num_l3,  'LIT': lit}

    predict_btn = st.button("🔍 Prediksi Kesiapan Olimpiade", use_container_width=True)

    if predict_btn:
        label, proba, total = predict_one(scores)
        status_name = CLASS_MAP[label]

        st.markdown("<hr class='divider'>", unsafe_allow_html=True)

        # Status display
        st.markdown("<div class='section-label'>HASIL PREDIKSI</div>", unsafe_allow_html=True)
        res_col1, res_col2 = st.columns([1, 2])

        with res_col1:
            css_class = {2: 'status-siap', 1: 'status-potensial', 0: 'status-tidak'}[label]
            prefix = nama_siswa + " — " if nama_siswa else ""
            st.markdown(f"""
            <div style='text-align:center; padding:1.2rem 0;'>
              <div style='font-size:0.75rem; color:rgba(255,255,255,0.4); text-transform:uppercase; letter-spacing:1.5px; margin-bottom:0.8rem;'>
                {prefix}Status Kesiapan
              </div>
              <div class='{css_class}'>
                {STATUS_EMOJI[label]} {status_name}
              </div>
              <div style='margin-top:1rem; font-size:0.85rem; color:rgba(255,255,255,0.5);'>
                Skor Total: <span style='color:#63b3ed; font-weight:700;'>{total:.2f}/100</span>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # Probability bars
            st.markdown("<div style='margin-top:0.8rem;'>", unsafe_allow_html=True)
            bar_colors = ['prob-fill-red', 'prob-fill-yellow', 'prob-fill-green']
            bar_labels  = ['Tidak Siap', 'Potensial', 'Siap Olimpiade']
            for i, (bl, bc) in enumerate(zip(bar_labels, bar_colors)):
                pct = proba[i] * 100
                st.markdown(f"""
                <div class='prob-row'>
                  <div class='prob-label'>{bl}</div>
                  <div class='prob-bar-bg'><div class='{bc}' style='width:{pct:.0f}%;'></div></div>
                  <div class='prob-pct'>{pct:.0f}%</div>
                </div>""", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with res_col2:
            st.plotly_chart(radar_chart(scores), use_container_width=True)

        st.plotly_chart(bar_chart(scores), use_container_width=True)

        # Review
        st.markdown("<div class='section-label'>REVIEW</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='review-box'>{get_review(label, scores, total)}</div>", unsafe_allow_html=True)

        # Rekomendasi
        st.markdown("<div class='section-label'>REKOMENDASI UNTUK GURU</div>", unsafe_allow_html=True)
        for rec in get_recommendations(label, scores):
            st.markdown(f"<div class='saran-item'>{rec}</div>", unsafe_allow_html=True)

        # Saran per komponen
        st.markdown("<div class='section-label'>SARAN PER KOMPONEN</div>", unsafe_allow_html=True)
        for s in get_saran(label, scores):
            st.markdown(f"<div class='saran-item'>{s}</div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════
#  TAB 2 — UPLOAD FILE EXCEL
# ═══════════════════════════════════════════════════════════
with tab2:
    st.markdown("<div class='card-title' style='margin-top:0.5rem;'>📂 Upload File Data Siswa</div>", unsafe_allow_html=True)

    st.markdown("""
    <div style='background:rgba(99,179,237,0.06); border:1px dashed rgba(99,179,237,0.3); border-radius:12px; padding:1rem 1.5rem; margin-bottom:1rem;'>
      <div style='font-size:0.85rem; color:rgba(255,255,255,0.7);'>
        📌 <b>Format file Excel yang diharapkan:</b><br>
        <span style='color:rgba(255,255,255,0.5);'>File harus memiliki kolom:</span>
        <code style='background:rgba(99,179,237,0.15); color:#63b3ed; padding:0.1rem 0.4rem; border-radius:4px; font-size:0.8rem;'>
          NUM_ALJ | NUM_GEO | NUM_BIL | NUM_DAT | NUM_L3 | LIT
        </code>
        <br><span style='color:rgba(255,255,255,0.4); font-size:0.8rem;'>Kolom Nama (opsional) dapat ditambahkan sebagai kolom pertama.</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader("Upload file Excel (.xlsx)", type=['xlsx', 'xls'])

    if uploaded:
        try:
            df_up = pd.read_excel(uploaded)
            st.success(f"✅ File berhasil dimuat: **{len(df_up)} baris data**")
            st.markdown("<div class='section-label'>PREVIEW DATA</div>", unsafe_allow_html=True)
            st.dataframe(df_up.head(5), use_container_width=True)

            # Deteksi kolom nama
            has_name = False
            name_col = None
            for c in df_up.columns:
                if c.strip().lower() in ['nama', 'name', 'nama siswa', 'student_name']:
                    has_name = True
                    name_col = c
                    break

            # Cek kolom numerik
            missing_cols = [c for c in FEATURE_COLS if c not in df_up.columns]
            if missing_cols:
                st.error(f"❌ Kolom berikut tidak ditemukan: {missing_cols}")
                st.stop()

            process_btn = st.button("⚡ Proses Semua Data & Prediksi", use_container_width=True)

            if process_btn:
                df_clean = df_up[FEATURE_COLS].dropna()
                valid_idx = df_up[FEATURE_COLS].dropna().index

                X = df_clean[FEATURE_COLS]
                X_sc = scaler.transform(X)
                labels  = rf_model.predict(X_sc)
                probas  = rf_model.predict_proba(X_sc)
                totals  = X.apply(lambda row: sum(row[c] * WEIGHTS[c] for c in FEATURE_COLS), axis=1)

                df_result = df_up.loc[valid_idx].copy()
                df_result['Total Score'] = totals.values.round(2)
                df_result['Status']      = [CLASS_MAP[l] for l in labels]
                df_result['Prob Tidak Siap (%)']    = (probas[:, 0] * 100).round(1)
                df_result['Prob Potensial (%)']      = (probas[:, 1] * 100).round(1)
                df_result['Prob Siap Olimpiade (%)'] = (probas[:, 2] * 100).round(1)
                df_result['Review'] = [get_review(l, dict(zip(FEATURE_COLS, X.iloc[i])), totals.iloc[i])
                                       for i, l in enumerate(labels)]

                st.markdown("<hr class='divider'>", unsafe_allow_html=True)
                st.markdown("<div class='section-label'>RINGKASAN DISTRIBUSI STATUS</div>", unsafe_allow_html=True)

                # Summary metrics
                counts = pd.Series(labels).value_counts()
                total_valid = len(labels)
                m1, m2, m3 = st.columns(3)
                m1.metric("🏆 Siap Olimpiade", f"{counts.get(2, 0)}", f"{counts.get(2, 0)/total_valid*100:.1f}%")
                m2.metric("⚡ Potensial",      f"{counts.get(1, 0)}", f"{counts.get(1, 0)/total_valid*100:.1f}%")
                m3.metric("❌ Tidak Siap",      f"{counts.get(0, 0)}", f"{counts.get(0, 0)/total_valid*100:.1f}%")

                # Pie chart
                pie_labels  = ['Siap Olimpiade', 'Potensial', 'Tidak Siap']
                pie_values  = [counts.get(2, 0), counts.get(1, 0), counts.get(0, 0)]
                pie_colors  = ['#10b981', '#f59e0b', '#ef4444']
                fig_pie = go.Figure(go.Pie(
                    labels=pie_labels, values=pie_values,
                    marker_colors=pie_colors,
                    hole=0.5,
                    textinfo='label+percent',
                    textfont=dict(color='white', size=11)
                ))
                fig_pie.update_layout(
                    paper_bgcolor='rgba(0,0,0,0)',
                    showlegend=False,
                    margin=dict(l=20, r=20, t=20, b=20),
                    height=280
                )

                c1, c2 = st.columns([1, 2])
                with c1:
                    st.plotly_chart(fig_pie, use_container_width=True)
                with c2:
                    avg_scores = {c: df_clean[c].mean() for c in FEATURE_COLS}
                    st.plotly_chart(bar_chart(avg_scores), use_container_width=True)

                st.markdown("<div class='section-label'>TABEL HASIL PREDIKSI</div>", unsafe_allow_html=True)
                st.dataframe(df_result, use_container_width=True, height=350)

                # Download
                excel_bytes = to_excel_download(df_result)
                st.download_button(
                    label="⬇️ Download Hasil (Excel)",
                    data=excel_bytes,
                    file_name="hasil_seleksi_olimpiade.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        except Exception as e:
            st.error(f"❌ Terjadi kesalahan saat memproses file: {e}")


# ═══════════════════════════════════════════════════════════
#  TAB 3 — INFO MODEL
# ═══════════════════════════════════════════════════════════
with tab3:
    st.markdown("<div class='card-title' style='margin-top:0.5rem;'>📈 Informasi & Performa Model</div>", unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Algoritma",    "Random Forest")
    c2.metric("Akurasi Test", f"{meta['accuracy']*100:.2f}%")
    c3.metric("CV Mean",      f"{meta['cv_mean']*100:.2f}%")
    c4.metric("CV Std",       f"±{meta['cv_std']*100:.2f}%")

    st.markdown("<div class='section-label'>FEATURE IMPORTANCE</div>", unsafe_allow_html=True)
    fi_sorted = sorted(meta['feature_importance'].items(), key=lambda x: -x[1])
    fi_labels = [FEATURE_LABELS[k] for k, _ in fi_sorted]
    fi_vals   = [v * 100 for _, v in fi_sorted]

    fig_fi = go.Figure(go.Bar(
        x=fi_vals, y=fi_labels, orientation='h',
        marker=dict(color=['#63b3ed'] * len(fi_vals),
                    line=dict(width=0)),
        text=[f"{v:.1f}%" for v in fi_vals],
        textposition='outside',
        textfont=dict(color='rgba(255,255,255,0.7)', size=11)
    ))
    fig_fi.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(title="Importance (%)", tickfont=dict(color='rgba(255,255,255,0.4)', size=9),
                   gridcolor='rgba(255,255,255,0.05)', range=[0, max(fi_vals) * 1.15]),
        yaxis=dict(tickfont=dict(color='rgba(255,255,255,0.7)', size=10), gridcolor='rgba(0,0,0,0)'),
        margin=dict(l=20, r=80, t=20, b=40),
        height=280
    )
    st.plotly_chart(fig_fi, use_container_width=True)

    st.markdown("<div class='section-label'>AMBANG BATAS KLASIFIKASI</div>", unsafe_allow_html=True)
    st.markdown(f"""
    <div style='display:flex; gap:1rem; flex-wrap:wrap; margin-top:0.5rem;'>
      <div class='metric-card'>
        <div class='metric-label'>Tidak Siap</div>
        <div class='metric-value' style='color:#ef4444; font-size:1rem;'>Skor &lt; {thresholds['p33']:.2f}</div>
      </div>
      <div class='metric-card'>
        <div class='metric-label'>Potensial</div>
        <div class='metric-value' style='color:#f59e0b; font-size:1rem;'>{thresholds['p33']:.2f} – {thresholds['p66']:.2f}</div>
      </div>
      <div class='metric-card'>
        <div class='metric-label'>Siap Olimpiade</div>
        <div class='metric-value' style='color:#10b981; font-size:1rem;'>Skor ≥ {thresholds['p66']:.2f}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div class='section-label'>BOBOT FITUR</div>", unsafe_allow_html=True)
    df_weights = pd.DataFrame([
        {'Fitur': FEATURE_LABELS[k], 'Kolom': k, 'Bobot': f"{v*100:.0f}%"}
        for k, v in WEIGHTS.items()
    ])
    st.dataframe(df_weights, use_container_width=True, hide_index=True)
